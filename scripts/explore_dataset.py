"""
DARE RAN Dataset — Dataset Exploration Script
=============================================
Run this script locally in PyCharm BEFORE any preprocessing or pipeline work.

What it produces (8 reports in ./reports/):
  01_dataset_overview.txt      — counts, structure, experiment design summary
  02_measurands_catalogue.csv  — every measurand: layer, priority, key_type, bounds
  03_feature_map.csv           — every x-feature mapped to measurand metadata
  04_layer_breakdown.txt       — measurands grouped by LTE protocol layer
  05_priority_features.txt     — priority-3 features only (highest importance)
  06_paper_benchmarks.txt      — x15, x42, x148, x182 fully explained
  07_domain_kpi_candidates.txt — 5 domain KPIs with source features mapped
  08_data_quality_rules.txt    — 8 Great Expectations rules derived from XLSX

CONFIRMED FOLDER STRUCTURE (from Folder_structure.txt):
  E:\\DARE_Data_Public\\
  └── Full_dataset\\
      ├── Tranche_A\\               <- 8 sessions, variable run counts (43-184)
      │   ├── Network_Protocol_Analyzer\\
      │   ├── Session_1\\
      │   │   ├── cipheron_0001\\
      │   │   │   └── traffic_generator\\   <- XLSX files live here
      │   │   ├── cipheroff_0003\\
      │   │   │   └── traffic_generator\\
      │   │   └── ...
      ├── Tranche_B\\               <- 13 sessions x exactly 60 runs each
      └── Tranche_C\\               <- 7  sessions x exactly 60 runs each

  Total confirmed: 2033 run folders, ~50,825 XLSX files across all tranches.

HOW TO RUN IN PYCHARM:
  Run -> Edit Configurations -> + -> Python
    Script path  : explore_dataset.py
    Parameters   : --dare_root "E:\\DARE_Data_Public\\Full_dataset"
                   --var_key "E:\\DARE_Data_Public\\variable_name_key.csv"
                   --measurands "E:\\DARE_Data_Public\\Measurands_and_Measurement_Terms.xlsx"
                   --output_dir "./reports"
    Working dir  : <your PyCharm project root, e.g. C:\\...\\dare-ran-pipeline>

  Omit --dare_root to skip the live folder scan (all 8 reports still run).

Requirements:
  pip install pandas openpyxl
"""

import argparse
import csv
import re
import sys
import textwrap
from collections import Counter, defaultdict
from pathlib import Path

try:
    import pandas as pd
    HAS_PANDAS = True
except ImportError:
    HAS_PANDAS = False

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed.  Run:  pip install openpyxl")
    sys.exit(1)

# ── Formatting ─────────────────────────────────────────────────────────────────
SECTION_LINE = "=" * 80
SUB_LINE     = "-" * 80


# ── LTE layer groupings ────────────────────────────────────────────────────────
LAYER_GROUPS = {
    "L1 Physical": [
        "Physical Layer", "Physical Layer (per UE)",
        "Physical Layer (Overview of the Cell)",
        "Physical Layer (Carrier Layer Overview of the Cell)",
        "Physical Layer (Channel Quality Indicator)",
        "Physical Layer (Hybrid Automatic Repeat Request)",
        "Physical Layer (Broadcast Channel)",
        "Physical Layer (Overview at the Cell)",
        "Physical Layer (Overview at the Cell for UL)",
        "Physical Layer (Channel Coding Summary Measurement)",
        "Physical Layer (UL SCH Transmission)",
        "Physical Layer (UL-SCH Hybrid Automatic Repeat Request)",
        "Physical Layer (UL Hybrid Automatic Repeat Request)",
        "Physical Layer (Physical Random Access Channel)",
        "Physical Layer (UL Sounding Reference Signal Transmission)",
        "Physical Layer (Duplication?)",
        "DL Control Channel (PDCCH, PCFICH, PHICH)",
    ],
    "L2 MAC":  ["MAC Layer", "MAC Layer (Overview of the Cell)",
                "MAC Layer (Overview of the Cell for UL)"],
    "L2 RLC":  ["RLC Layers"],
    "L2 PDCP": ["PDCP Layer (Overview of the Cell)",
                "DL PDCP Layer (Overview of the Cell)",
                "UL PDCP Layer (Overview of the Cell)"],
    "L3/App/System": ["Application Layer", "System Layers (Overview of the Cell)"],
    "Wireshark":     ["Network"],
}

# ── Paper benchmarks ───────────────────────────────────────────────────────────
PAPER_BENCHMARKS = {
    "x15": {
        "orig_name": "l1carrierpowers_real_rsrq__pcc_db-r_-10.5",
        "meaning":   "Proportion of RSRQ readings equal to -10.5 dB",
        "expected":  "approx 0.34 for Tranche_B / Session_6 / run_317 (cipher ON)",
        "paper_ref": "Section V-C, Figure 8",
        "layer":     "L1 Physical",
        "why_important": (
            "RSRQ (Reference Signal Received Quality) measures signal quality. "
            "The proportion at exactly -10.5 dB separates cipher-on from cipher-off "
            "because ciphering affects RLC/MAC overhead which back-propagates to "
            "per-subframe resource allocation and thus the RSRQ distribution shape."
        ),
    },
    "x42": {
        "orig_name": "l1dlcarrierstats_dl_sch_bler-q_0.95",
        "meaning":   "95th percentile of DL-SCH Block Error Rate",
        "expected":  "Small but consistent separation between cipher states (Figure 9)",
        "paper_ref": "Section V-C, Figure 9",
        "layer":     "L1 Physical",
        "why_important": (
            "BLER (Block Error Rate) is the ratio of CRC failures to total blocks. "
            "The 0.95 quantile captures tail behaviour. Cipher-off runs show slightly "
            "higher BLER tails due to changed MAC scheduling dynamics."
        ),
    },
    "x148": {
        "orig_name": "l1dlcarrierstats_retransmission_1-r_0.0",
        "meaning":   "Proportion of observations where retransmission count = 0",
        "expected":  "approx 0.925-0.945; cipher-on slightly higher (Figure 10)",
        "paper_ref": "Section V-C, Figure 10",
        "layer":     "L1 Physical",
        "why_important": (
            "Retransmission count = 0 means first-time delivery success. Cipher-off "
            "runs show marginally fewer zero-retransmission events, suggesting the "
            "absence of encryption overhead subtly changes HARQ timing."
        ),
    },
    "x182": {
        "orig_name": "macrxstats_mean_pdu_size_bytes-q_0.25",
        "meaning":   "25th percentile of mean MAC PDU size in bytes",
        "expected":  "approx 268.2-268.7 bytes (Figure 11)",
        "paper_ref": "Section V-C, Figure 11",
        "layer":     "L2 MAC",
        "why_important": (
            "MAC PDU size reflects data packing per transmission unit. Encryption "
            "adds fixed overhead per PDU. With cipher disabled PDU sizes shift "
            "measurably at the distribution tails."
        ),
    },
}

# ── Domain KPI definitions ─────────────────────────────────────────────────────
DOMAIN_KPIS = {
    "harq_efficiency": {
        "formula":    "ACK count / (ACK count + NACK count) per run",
        "source":     ["l1dlcarrierstats_average_retransmission_count (x35-x37)",
                       "DLHARQRX ACK/NACK: Enum 0=NACK 1=ACK 2=N/A"],
        "x_features": ["x35", "x36", "x37", "x148"],
        "layer":      "L1 Physical (HARQ)",
        "range":      "[0.0, 1.0]",
        "signal": (
            "Retransmission load. Cipher-off removes the encryption step from "
            "the RLC/MAC-to-PDCP handoff, slightly changing ACK/NACK timing. "
            "Efficiency expected marginally higher cipher-off due to reduced "
            "per-PDU processing latency at the UE."
        ),
    },
    "spectral_efficiency": {
        "formula":    "Aggregate DL-SCH throughput (kbps) / channel bandwidth (MHz)",
        "source":     ["l1celldloverview_aggregate_pcc_dl_sch_throughput_kbps (x22-x26)",
                       "Bandwidth: fixed 10 MHz in DARE testbed (paper Section II)"],
        "x_features": ["x22", "x23", "x24", "x25", "x26"],
        "layer":      "L1 Physical (Cell Overview)",
        "range":      "Approx [0, 30000] for this testbed configuration",
        "signal": (
            "Cipher-off removes ~8 bytes of AES overhead per PDU. At 15 UEs "
            "this aggregate difference produces a small but consistent throughput "
            "uplift visible at the median and upper quantiles."
        ),
    },
    "bler_spread": {
        "formula":    "q0.95(dl_sch_bler) minus q0.05(dl_sch_bler) per run",
        "source":     ["l1dlcarrierstats_dl_sch_bler-q_0.05 (x38)",
                       "l1dlcarrierstats_dl_sch_bler-q_0.95 (x42)"],
        "x_features": ["x38", "x42"],
        "layer":      "L1 Physical (DL Carrier Stats)",
        "range":      "[0.0, 1.0]  (difference of two [0,1] values)",
        "signal": (
            "Distribution width of block error rate. Wide spread = RLC instability "
            "within the run. Expected to differ systematically between cipher states "
            "as MAC scheduling dynamics change."
        ),
    },
    "dl_ul_asymmetry": {
        "formula":    "median(DL-SCH throughput kbps) / median(UL-SCH throughput kbps)",
        "source":     ["l1dlcarrierstats_dl_sch_throughput_kbps-q_0.5 (x45)",
                       "rlctxstats_all_sdu_packet_rate-q_0.5 (x405-x407 area)"],
        "x_features": ["x45", "x405", "x406", "x407"],
        "layer":      "L2 MAC / L2 RLC",
        "range":      "Typically [2, 20] for LTE FDD DL-heavy traffic",
        "signal": (
            "LTE FDD cipher overhead is applied symmetrically in UL and DL at PDCP. "
            "Any shift in asymmetry ratio indicates cipher-off affects one direction "
            "more than the other -- a useful discriminating signal."
        ),
    },
    "pdcp_discard_rate": {
        "formula":    "macrxstats_discards proportion / total PDUs per run",
        "source":     ["macrxstats_discards proportion features",
                       "macrxstats_number_of_pdus quantile features",
                       "Check pdcprxstats_* features for direct PDCP counters"],
        "x_features": ["resolve from pdcprxstats_ prefix in feature_map.csv"],
        "layer":      "L2 PDCP  <-- HIGHEST PRIORITY: cipher executes at this layer",
        "range":      "[0.0, 1.0]",
        "signal": (
            "PDCP is where AES encryption/decryption physically occurs in LTE. "
            "With cipher OFF the PDCP layer no longer drops PDUs with failed "
            "integrity checks -- discard rate should drop. This is the most direct "
            "cipher-state proxy in the entire feature set."
        ),
    },
}


# ── Utilities ──────────────────────────────────────────────────────────────────
def section(title):
    print(f"\n{SECTION_LINE}\n  {title}\n{SECTION_LINE}")


def write_report(path, content):
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(content, encoding="utf-8")
    kb = path.stat().st_size // 1024
    print(f"  -> {path.name}  ({kb} KB)")


def wrap(text, width=74, indent="    "):
    return "\n".join(
        textwrap.fill(line, width=width,
                      initial_indent=indent, subsequent_indent=indent)
        for line in text.splitlines() if line.strip()
    )


def assign_layer_group(layer_raw):
    if not layer_raw:
        return "Unknown"
    for group, layers in LAYER_GROUPS.items():
        for lyr in layers:
            if lyr.lower() in layer_raw.lower():
                return group
    return "Other"


def parse_feature_name(orig):
    m = re.match(r"^(.+)-q_([0-9.]+)$", orig)
    if m:
        return {"measurand_prefix": m.group(1), "feature_type": "quantile",
                "feature_value": float(m.group(2))}
    m = re.match(r"^(.+)-r_(.+)$", orig)
    if m:
        return {"measurand_prefix": m.group(1), "feature_type": "proportion",
                "feature_value": m.group(2)}
    return {"measurand_prefix": orig, "feature_type": "metadata",
            "feature_value": None}


# ── Loaders ────────────────────────────────────────────────────────────────────
def load_variable_key(path):
    with open(path, newline="", encoding="utf-8") as f:
        rows = list(csv.DictReader(f))
    return [{k: v.strip() for k, v in r.items()} for r in rows]


def load_measurands(path):
    wb = openpyxl.load_workbook(path)

    def to_dicts(ws):
        headers = [c.value for c in ws[1]]
        return [
            {headers[i]: row[i] for i in range(len(headers)) if headers[i]}
            for row in ws.iter_rows(min_row=2, values_only=True)
            if any(v is not None for v in row[:len(headers)])
        ]

    return {
        "utg":  to_dicts(wb["UTG Measurands"]),
        "defs": to_dicts(wb["UTG Definitions"]),
        "net":  to_dicts(wb["Network Protocol Measurands"]),
    }


def scan_dare_root(dare_root):
    """
    Walk Full_dataset and count runs per session.
    Reads folder names only -- no XLSX files are opened.

    Structure:  dare_root / Tranche_A / Session_1 / cipheron_0001 / traffic_generator
    """
    if not dare_root.exists():
        return {}

    stats = {}
    tranche_dirs = sorted(
        d for d in dare_root.iterdir()
        if d.is_dir() and d.name.startswith("Tranche_")
    )
    for t_dir in tranche_dirs:
        tranche = t_dir.name
        session_dirs = sorted(
            d for d in t_dir.iterdir()
            if d.is_dir() and d.name.startswith("Session_")
        )
        for s_dir in session_dirs:
            key = f"{tranche}/{s_dir.name}"
            stats[key] = {"on": 0, "off": 0, "xlsx_groups": set()}

            for r_dir in s_dir.iterdir():
                if not r_dir.is_dir():
                    continue
                name = r_dir.name.lower()
                if "cipheron" in name:
                    stats[key]["on"] += 1
                elif "cipheroff" in name:
                    stats[key]["off"] += 1

                # Peek at one traffic_generator to learn XLSX group names
                tg = r_dir / "traffic_generator"
                if tg.exists() and not stats[key]["xlsx_groups"]:
                    stats[key]["xlsx_groups"] = {
                        f.name for f in tg.iterdir()
                        if f.suffix.lower() == ".xlsx"
                    }
    return stats


# ── Reports ────────────────────────────────────────────────────────────────────
def report_01(var_rows, meas, dare_stats, output_dir):
    section("REPORT 1: Dataset Overview")

    utg  = meas["utg"]
    DCOL = "Decision (0 - dont use, 1 - 25% use, 2 - 50% use, 3 - 100% use)"
    KCOL = "Key types: Continuous, discrete, unknown, and NA"

    feat_types = Counter(
        parse_feature_name(r["orig_var_name"])["feature_type"] for r in var_rows
    )
    decisions = Counter(r.get(DCOL) for r in utg)
    key_types = Counter(r.get(KCOL) for r in utg)

    lines = [
        SECTION_LINE, "DARE DATASET -- OVERVIEW REPORT", SECTION_LINE, "",
        "SOURCE FILES", SUB_LINE,
        f"  variable_name_key.csv      : {len(var_rows)} features  (x1 - x{len(var_rows)})",
        f"  Measurands XLSX / UTG      : {len(utg)} measurand rows",
        f"  Measurands XLSX / Network  : {len(meas['net'])} network protocol measurands",
        f"  Measurands XLSX / Defs     : {len(meas['defs'])} abbreviation definitions",
        "",
        "FEATURE BREAKDOWN  (variable_name_key.csv)", SUB_LINE,
        f"  Total x-features                : {len(var_rows)}",
        f"  Quantile features  (-q_)        : {feat_types['quantile']}",
        f"    |_ continuous measurands at q0.05 / q0.25 / q0.50 / q0.75 / q0.95",
        f"  Proportion features (-r_)       : {feat_types['proportion']}",
        f"    |_ discrete measurands as fraction of observations at each value",
        f"  Metadata / other (x1-x3, VSA)  : {feat_types['metadata']}",
        "",
        "MEASURAND PRIORITY  (UTG Measurands sheet)", SUB_LINE,
        f"  Priority 3  use 100%  : {decisions.get(3, 0):>4}  <- your primary feature set",
        f"  Priority 2  use ~50%  : {decisions.get(2, 0):>4}",
        f"  Priority 1  use ~25%  : {decisions.get(1, 0):>4}",
        f"  Priority 0  exclude   : {decisions.get(0, 0):>4}",
        f"  Total                 : {sum(decisions.values()):>4}",
        "",
        "MEASURAND KEY TYPE", SUB_LINE,
        f"  Continuous (q=5)  : {key_types.get('continuous (q = 5)', 0):>4}",
        f"  Continuous (q=3)  : {key_types.get('continuous (q = 3)', 0):>4}",
        f"  Discrete          : {key_types.get('discrete', 0):>4}  (-> proportion features)",
        f"  NA (excluded)     : {key_types.get('NA', 0):>4}",
        "",
        "EXPERIMENT DESIGN  (from DARE paper)", SUB_LINE,
        "  Tranches       : 3  (A = Sep 2022 | B = Jan 2023 | C = Apr 2023)",
        "  Sessions       : 8 (Tranche A) | 13 (Tranche B) | 7 (Tranche C)",
        "  Runs/session   : Variable 43-184 (Tranche A) | Exactly 60 (B and C)",
        "  Run duration   : 5 min measurement + 11 min reset = 16 min total",
        "  Cipher pattern : on / on / off / off  (4-run repeating cycle per session)",
        "  UEs            : 15 simulated user equipments per run",
        "  Sample rate    : 200 ms intervals -> ~1 500 rows per measurand per run",
        "  Time trim      : first and last 5 seconds removed per run (paper IV-B)",
        "  ML split       : Tranche B = training | Tranches A and C = validation",
    ]

    if dare_stats:
        lines += [
            "",
            "CONFIRMED RUN COUNTS  (live scan of your Full_dataset folder)", SUB_LINE,
            f"  {'Session':<28} {'CipherON':>10} {'CipherOFF':>10} {'Total':>8}",
            "  " + "-" * 60,
        ]
        tranche_totals = {}
        for sess, cnt in sorted(dare_stats.items()):
            on, off, total = cnt["on"], cnt["off"], cnt["on"] + cnt["off"]
            lines.append(f"  {sess:<28} {on:>10} {off:>10} {total:>8}")
            tranche = sess.split("/")[0]
            if tranche not in tranche_totals:
                tranche_totals[tranche] = {"on": 0, "off": 0, "sessions": 0}
            tranche_totals[tranche]["on"]       += on
            tranche_totals[tranche]["off"]      += off
            tranche_totals[tranche]["sessions"] += 1

        lines += [
            "",
            f"  {'Tranche':<28} {'CipherON':>10} {'CipherOFF':>10} {'Total':>8}  Sessions",
            "  " + "-" * 68,
        ]
        grand_on = grand_off = 0
        for tranche, t in sorted(tranche_totals.items()):
            tot = t["on"] + t["off"]
            lines.append(
                f"  {tranche:<28} {t['on']:>10} {t['off']:>10} {tot:>8}"
                f"  ({t['sessions']} sessions)"
            )
            grand_on  += t["on"]
            grand_off += t["off"]
        grand = grand_on + grand_off
        lines += [
            f"  {'GRAND TOTAL':<28} {grand_on:>10} {grand_off:>10} {grand:>8}",
            "",
            f"  Estimated XLSX files (25 groups x {grand} runs) : ~{grand * 25:,}",
        ]

        # Show XLSX group names from first run found
        for cnt in dare_stats.values():
            if cnt["xlsx_groups"]:
                lines += [
                    "",
                    f"  XLSX groups found in one traffic_generator folder"
                    f"  ({len(cnt['xlsx_groups'])} files):",
                ]
                for g in sorted(cnt["xlsx_groups"]):
                    lines.append(f"    * {g}")
                break

    content = "\n".join(lines)
    print(content)
    write_report(output_dir / "01_dataset_overview.txt", content)


def report_02(meas, output_dir):
    section("REPORT 2: Measurands Catalogue")

    DCOL = "Decision (0 - dont use, 1 - 25% use, 2 - 50% use, 3 - 100% use)"
    KCOL = "Key types: Continuous, discrete, unknown, and NA"

    rows_out = [
        {
            "layer_group": assign_layer_group(r.get("Layer")),
            "layer_raw":   r.get("Layer", ""),
            "ue_logging":  r.get("UE Logging", ""),
            "meas_key":    r.get("Meas. Item Key (for Python coding)", ""),
            "priority":    r.get(DCOL, ""),
            "key_type":    r.get(KCOL, ""),
            "data_type":   r.get("Type", ""),
            "min_value":   r.get("Minimum Value", ""),
            "max_value":   r.get("Maximum Value", ""),
            "description": r.get("Description", ""),
        }
        for r in meas["utg"]
    ]

    out = output_dir / "02_measurands_catalogue.csv"
    if HAS_PANDAS:
        df = pd.DataFrame(rows_out)
        df.to_csv(out, index=False)
        print(f"  {len(df)} measurand rows written")
        print("\n  Layer group distribution:")
        print(df.groupby("layer_group")["meas_key"].count()
                .sort_values(ascending=False).to_string())
        print("\n  Priority x key_type cross-tab:")
        print(df.groupby(["priority", "key_type"]).size()
                .unstack(fill_value=0).to_string())
    else:
        with open(out, "w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=rows_out[0].keys())
            w.writeheader()
            w.writerows(rows_out)
        print(f"  {len(rows_out)} rows written")


def report_03(var_rows, meas, output_dir):
    section("REPORT 3: Feature Map  (x1-x424 with metadata)")

    DCOL = "Decision (0 - dont use, 1 - 25% use, 2 - 50% use, 3 - 100% use)"
    KCOL = "Key types: Continuous, discrete, unknown, and NA"

    meas_idx = defaultdict(list)
    for r in meas["utg"]:
        ue = (r.get("UE Logging") or "").lower().replace("_", "")
        if ue:
            meas_idx[ue].append(r)

    rows_out = []
    for vr in var_rows:
        fp  = parse_feature_name(vr["orig_var_name"])
        pfx = fp["measurand_prefix"].lower().replace("_", "")

        matched = None
        for ue_key, entries in meas_idx.items():
            if ue_key in pfx or pfx in ue_key:
                matched = entries[0]
                break

        rows_out.append({
            "shorthand":        vr["new_var_num"],
            "orig_var_name":    vr["orig_var_name"],
            "measurand_prefix": fp["measurand_prefix"],
            "feature_type":     fp["feature_type"],
            "feature_value":    fp["feature_value"],
            "layer_group":      assign_layer_group(matched.get("Layer") if matched else None),
            "layer_raw":        (matched.get("Layer") or "") if matched else "",
            "priority":         (matched.get(DCOL) or "") if matched else "",
            "key_type":         (matched.get(KCOL) or "") if matched else "",
            "data_type":        (matched.get("Type") or "") if matched else "",
            "min_value":        (matched.get("Minimum Value") or "") if matched else "",
            "max_value":        (matched.get("Maximum Value") or "") if matched else "",
        })

    out = output_dir / "03_feature_map.csv"
    if HAS_PANDAS:
        df = pd.DataFrame(rows_out)
        df.to_csv(out, index=False)
        print(f"  {len(df)} feature rows written")
        print("\n  Features by type:")
        print(df["feature_type"].value_counts().to_string())
    else:
        with open(out, "w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=rows_out[0].keys())
            w.writeheader()
            w.writerows(rows_out)
        print(f"  {len(rows_out)} rows written")


def report_04(var_rows, meas, output_dir):
    section("REPORT 4: LTE Protocol Layer Breakdown")

    DCOL = "Decision (0 - dont use, 1 - 25% use, 2 - 50% use, 3 - 100% use)"

    layer_priority = defaultdict(Counter)
    layer_p3_meas  = defaultdict(list)

    for r in meas["utg"]:
        lg  = assign_layer_group(r.get("Layer"))
        pri = r.get(DCOL)
        layer_priority[lg][pri] += 1
        if pri == 3:
            m = r.get("Meas. Item Key (for Python coding)", "")
            d = (r.get("Description") or "")[:90]
            if m:
                layer_p3_meas[lg].append((m, d))

    prefix_layer = {
        "l1": "L1 Physical", "vsa": "L1 Physical",
        "mac": "L2 MAC", "rlc": "L2 RLC",
        "pdcp": "L2 PDCP", "sys": "L3/App/System",
        "wireshark": "Wireshark",
    }
    feat_layer_cnt = Counter()
    for vr in var_rows:
        orig = vr["orig_var_name"].lower()
        matched = "Metadata/Other"
        for pref, lg in prefix_layer.items():
            if orig.startswith(pref):
                matched = lg
                break
        feat_layer_cnt[matched] += 1

    layer_explanations = {
        "L1 Physical": (
            "Physical radio signal quality (RSRP, RSRQ, SNR), DL scheduling "
            "(MCS, resource block allocation), error rates (BLER), HARQ "
            "retransmission behaviour, and per-subframe throughput. "
            "Largest layer -- 130 of the 424 x-features originate here."
        ),
        "L2 MAC": (
            "Medium Access Control: PDU sizes, SDU packet rates, retransmission "
            "counts, padding overhead, control block counts. MAC coordinates "
            "scheduling and error correction. MAC PDU sizes directly reflect "
            "cipher overhead -- see benchmark feature x182."
        ),
        "L2 RLC": (
            "Radio Link Control: SDU throughput, packet rates, PDU counts for "
            "both DL (rx) and UL (tx). RLC sits between MAC and PDCP handling "
            "segmentation and reassembly."
        ),
        "L2 PDCP": (
            "Packet Data Convergence Protocol: SDU rates, throughput, discards. "
            "AES encryption/decryption executes physically at this layer in LTE. "
            "Highest-priority layer for cipher-state detection -- discard rate "
            "and throughput are most directly affected by cipher-off. "
            "pdcp_discard_rate is your most sensitive domain KPI."
        ),
        "L3/App/System": (
            "System overview aggregates and application layer statistics. "
            "Cell-level summaries not tied to a specific protocol sublayer."
        ),
        "Wireshark": (
            "S1-MME interface packet captures between eNB and EPC. "
            "7 features (x418-x424): total packets, packets/sec, eNB-as-source "
            "vs destination counts. Network control-plane view."
        ),
    }

    lines = [
        SECTION_LINE, "LTE PROTOCOL LAYER BREAKDOWN", SECTION_LINE, "",
        "Measurements span 27 LTE protocol sub-layers, consolidated into 5",
        "functional groups. Read this before touching any feature names.", "",
    ]

    for group in ["L1 Physical", "L2 MAC", "L2 RLC", "L2 PDCP",
                  "L3/App/System", "Wireshark"]:
        pc = layer_priority.get(group, Counter())
        fc = feat_layer_cnt.get(group, 0)
        p3 = layer_p3_meas.get(group, [])

        lines += [
            "-" * 70,
            f"  {group}",
            "-" * 70,
            f"  Measurands total   : {sum(pc.values())}",
            f"  Priority-3 count   : {pc.get(3, 0)}",
            f"  x-features derived : {fc}",
            "",
            "  What this layer captures:",
            wrap(layer_explanations.get(group, ""), width=72),
            "",
        ]
        if p3:
            lines.append("  Priority-3 measurands (first 6):")
            for m, d in p3[:6]:
                lines.append(f"    * {m}")
                if d:
                    lines.append(wrap(d, width=68, indent="        "))
        lines.append("")

    content = "\n".join(lines)
    print(content[:2000], "\n  ...[see 04_layer_breakdown.txt for full output]")
    write_report(output_dir / "04_layer_breakdown.txt", content)


def report_05(var_rows, meas, output_dir):
    section("REPORT 5: Priority-3 Features")

    DCOL = "Decision (0 - dont use, 1 - 25% use, 2 - 50% use, 3 - 100% use)"
    p3_ue_logs = {
        (r.get("UE Logging") or "").lower()
        for r in meas["utg"]
        if r.get(DCOL) == 3 and r.get("UE Logging")
    }

    matched = []
    for vr in var_rows:
        fp     = parse_feature_name(vr["orig_var_name"])
        prefix = fp["measurand_prefix"].lower()
        for ue in p3_ue_logs:
            if ue in prefix:
                matched.append({
                    "shorthand":    vr["new_var_num"],
                    "orig":         vr["orig_var_name"],
                    "feature_type": fp["feature_type"],
                    "value":        str(fp["feature_value"]),
                })
                break

    lines = [
        SECTION_LINE,
        f"PRIORITY-3 FEATURES  --  {len(matched)} x-features",
        SECTION_LINE, "",
        "These features correspond to Decision=3 measurands (highest importance).",
        "Use these first for EDA, feature importance, and initial ML runs.", "",
        f"  {'Shorthand':<10} {'Type':<12} {'Value':<12}  Orig Variable Name",
        "  " + "-" * 76,
    ]
    for f in sorted(matched, key=lambda x: int(x["shorthand"][1:])):
        lines.append(
            f"  {f['shorthand']:<10} {f['feature_type']:<12} {f['value']:<12}  {f['orig']}"
        )

    content = "\n".join(lines)
    print(f"  Priority-3 x-features matched: {len(matched)}")
    print(f"  Sample: {[f['shorthand'] for f in matched[:12]]}")
    write_report(output_dir / "05_priority_features.txt", content)


def report_06(output_dir):
    section("REPORT 6: Paper Benchmark Features")

    lines = [
        SECTION_LINE, "PAPER BENCHMARK FEATURES  (x15, x42, x148, x182)",
        SECTION_LINE, "",
        "Validated explicitly in the DARE Data Guide paper Section V-C (Figures 8-11).",
        "Compute these first after running preprocess.py to verify your pipeline.", "",
    ]

    for xid, info in PAPER_BENCHMARKS.items():
        lines += [
            "-" * 70,
            f"  {xid.upper()}  --  {info['orig_name']}",
            "-" * 70,
            f"  Meaning    : {info['meaning']}",
            f"  Expected   : {info['expected']}",
            f"  Paper ref  : {info['paper_ref']}",
            f"  Layer      : {info['layer']}", "",
            "  Why it matters for cipher-state detection:",
            wrap(info["why_important"], width=72), "",
        ]

    lines += [
        SECTION_LINE, "VALIDATION CHECKLIST", SECTION_LINE, "",
        "After preprocess.py + transform.py, verify these four values:", "",
        "  [] x15 for Tranche_B / Session_6 / run_317 (cipher=on)  approx 0.34",
        "     Proportion of RSRQ readings at exactly -10.5 dB", "",
        "  [] x42 across Tranche_B runs: cipher-on cluster slightly BELOW cipher-off",
        "     (lower 95th-pctile BLER when encrypted)", "",
        "  [] x148 values cluster in [0.925, 0.945]; cipher-on slightly higher", "",
        "  [] x182 values cluster in [268.2, 268.7] bytes;",
        "     cipher-off slightly lower (no encryption header overhead)", "",
        "If any value is significantly off, check:",
        "  1. Time-trimming applied (+-5 s from each end)?",
        "  2. Cipher labels cross-validated (folder name vs metadata.db)?",
        "  3. Proportion denominator = total observations AFTER trimming?",
        "  4. Run 317 is in Tranche_B / Session_6 -- correct session folder?",
    ]

    content = "\n".join(lines)
    print(content)
    write_report(output_dir / "06_paper_benchmarks.txt", content)


def report_07(var_rows, output_dir):
    section("REPORT 7: Domain KPI Candidates")

    xmap = {r["new_var_num"]: r["orig_var_name"] for r in var_rows}

    lines = [
        SECTION_LINE, "TELECOM DOMAIN KPIs  --  dare_gold.domain_features",
        SECTION_LINE, "",
        "5 engineered features computed in Phase 3 alongside x1-x424.",
        "These use your telecom expertise to create signals no generic data",
        "engineer can derive from the raw features alone.", "",
    ]

    for name, info in DOMAIN_KPIS.items():
        found = [
            (xid, xmap.get(xid, "NOT FOUND in variable_key"))
            for xid in info["x_features"]
            if not xid.startswith("resolve")
        ]

        lines += [
            "-" * 70,
            f"  {name.upper()}",
            "-" * 70,
            f"  Formula        : {info['formula']}",
            f"  Layer          : {info['layer']}",
            f"  Expected range : {info['range']}", "",
            "  Source measurands:",
        ]
        for src in info["source"]:
            lines.append(f"    * {src}")
        lines += ["", "  Mapped x-features:"]
        for xid, orig in found:
            ok = "OK" if "NOT FOUND" not in orig else "XX"
            lines.append(f"    [{ok}] {xid:<8} {orig}")
        lines += [
            "", "  Cipher-state sensitivity:",
            wrap(info["signal"], width=72), "",
        ]

    content = "\n".join(lines)
    print(content[:1500], "\n  ...[see 07_domain_kpi_candidates.txt for full output]")
    write_report(output_dir / "07_domain_kpi_candidates.txt", content)


def report_08(output_dir):
    section("REPORT 8: Data Quality Rules")

    rules = [
        ("DQ-01", "Cipher state completeness",
         "cipher_state column",
         "expect_column_values_to_be_in_set",
         '{"value_set": ["on", "off"]}',
         "cipher_state must be exactly 'on' or 'off' -- never null, never 'unknown'. "
         "Any 'unknown' means dual validation (folder name vs metadata.db) failed "
         "and the run must be quarantined, not silently loaded.",
         "Dual validation logic: folder name + metadata.db eNB column"),

        ("DQ-02", "RSRQ proportion sum <= 1.0",
         "x14 + x15 + x16  (RSRQ proportions at -10.4, -10.5, -10.6 dB)",
         "custom: sum(x14, x15, x16) per run <= 1.0",
         "sum <= 1.0",
         "x14/x15/x16 are parts of the RSRQ distribution -- their sum cannot exceed "
         "1.0. A sum > 1.0 means double-counting or a wrong denominator.",
         "Measurands XLSX: Real RSRQ (PCC)(dB) -- discrete distribution"),

        ("DQ-03", "RSRP range bounds",
         "x9-x13  (l1carrierpowers RSRP proportion features)",
         "expect_column_values_to_be_between",
         '{"min_value": -72.0, "max_value": -65.3}',
         "DARE testbed uses a fixed conductive attenuator -- RSRP stays in a "
         "narrow range. Values outside [-72.0, -65.3] dBm indicate an RF "
         "environment change or measurement artefact.",
         "Measurands XLSX: Real RSRP (PCC)(dBm) min=-72.0 max=-65.3"),

        ("DQ-04", "Row count floor per run",
         "row count after time-trimming",
         "expect_table_row_count_to_be_between",
         '{"min_value": 1000, "max_value": 20000}',
         "5 min at 200 ms = ~1500 rows. After +-5 s trim expect >= 1400. "
         "Under 1000 rows means a measurement failure. Upper bound 20000 covers "
         "run 317 which has 18824 rows (15 UEs x ~1250 obs each).",
         "DARE paper V-C: 'After removing 5 seconds... 18,824 data points'"),

        ("DQ-05", "Proportion features bounded [0, 1]",
         "all -r_ features",
         "expect_column_values_to_be_between",
         '{"min_value": 0.0, "max_value": 1.0}',
         "Proportion features are fractions -- mathematically cannot be negative "
         "or exceed 1.0. Any violation means the denominator calculation is wrong.",
         "Mathematical constraint on all proportion features"),

        ("DQ-06", "Quantile ordering constraint",
         "all -q_ features (same measurand, 5 quantiles)",
         "custom multi-column expectation",
         "q0.05 <= q0.25 <= q0.50 <= q0.75 <= q0.95  per measurand per run",
         "Quantiles must be non-decreasing. If q0.95 < q0.75 for the same "
         "measurand the computation has a bug. Example: x38 (q0.05 BLER) must "
         "always be <= x42 (q0.95 BLER) for every run.",
         "Mathematical constraint on order statistics"),

        ("DQ-07", "HARQ efficiency in [0, 1]",
         "harq_efficiency  (domain_features table)",
         "expect_column_values_to_be_between",
         '{"min_value": 0.0, "max_value": 1.0}',
         "harq_efficiency = ACK / (ACK + NACK). Must be in [0, 1] as a ratio. "
         "Values outside this range mean source proportion features are malformed "
         "(see DQ-05).",
         "Domain KPI definition: ACK rate / (ACK + NACK rate)"),

        ("DQ-08", "Pipeline run completeness",
         "pipeline_runs log vs run_manifest",
         "custom: COUNT comparison",
         "COUNT(pipeline_runs WHERE status='pass') == COUNT(run_manifest)",
         "Every run in dare_gold.run_manifest must have a passing pipeline_runs "
         "entry. Any run in the manifest but missing from pipeline_runs was "
         "silently skipped -- which must never happen.",
         "Pipeline design constraint: no silent failures"),
    ]

    lines = [
        SECTION_LINE, "DATA QUALITY RULES  --  8 Great Expectations Rules",
        SECTION_LINE, "",
        "Derived from Measurands XLSX bounds, data types, and domain constraints.",
        "Commit these as /configs/ge_expectations.json in your GitHub repo.", "",
    ]

    for rid, name, col, method, params, rationale, source in rules:
        lines += [
            "-" * 70,
            f"  {rid}  {name}",
            "-" * 70,
            f"  Column(s)   : {col}",
            f"  GE method   : {method}",
            f"  Parameters  : {params}",
            f"  Source      : {source}", "",
            "  Rationale:",
            wrap(rationale, width=72), "",
        ]

    lines += [
        SECTION_LINE, "QUICK-START SNIPPET", SECTION_LINE, "",
        "  from great_expectations.dataset import PandasDataset",
        "  ds = PandasDataset(feature_matrix_df)",
        "",
        "  # DQ-01",
        "  r = ds.expect_column_values_to_be_in_set('cipher_state', ['on', 'off'])",
        "  assert r.success, f'DQ-01 FAILED: {r}'",
        "",
        "  # DQ-05  (run for every proportion column)",
        "  for col in proportion_cols:",
        "      r = ds.expect_column_values_to_be_between(col, 0.0, 1.0)",
        "      assert r.success, f'DQ-05 FAILED on {col}'",
        "",
        "Store the GE HTML report in /reports/data_quality_report.html.",
    ]

    content = "\n".join(lines)
    print(f"  8 rules documented")
    for rid, name, *_ in rules:
        print(f"    {rid}: {name}")
    write_report(output_dir / "08_data_quality_rules.txt", content)


# ── Main ───────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(
        description="DARE Dataset Exploration -- 8 structured reports",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=textwrap.dedent("""
        PyCharm Parameters field (Run -> Edit Configurations):
          --dare_root  "E:\\DARE_Data_Public\\Full_dataset"
          --var_key    "E:\\DARE_Data_Public\\variable_name_key.csv"
          --measurands "E:\\DARE_Data_Public\\Measurands_and_Measurement_Terms.xlsx"
          --output_dir "./reports"

        Omit --dare_root to skip the live folder scan.
        """)
    )
    parser.add_argument("--var_key",     default="variable_name_key.csv")
    parser.add_argument("--measurands",  default="Measurands_and_Measurement_Terms.xlsx")
    parser.add_argument("--dare_root",   default=None,
                        help="Path to Full_dataset folder (optional, enables live scan)")
    parser.add_argument("--output_dir",  default="./reports")
    args = parser.parse_args()

    var_key_path    = Path(args.var_key)
    measurands_path = Path(args.measurands)
    dare_root       = Path(args.dare_root) if args.dare_root else None
    output_dir      = Path(args.output_dir)

    # Validate
    errors = []
    for p in [var_key_path, measurands_path]:
        if not p.exists():
            errors.append(f"File not found: {p}")
    if dare_root and not dare_root.exists():
        errors.append(f"dare_root not found: {dare_root}")
    if errors:
        for e in errors:
            print(f"ERROR: {e}")
        print("\nTip: use full absolute paths in --var_key / --measurands / --dare_root")
        sys.exit(1)

    output_dir.mkdir(parents=True, exist_ok=True)

    print(f"\n{'=' * 80}")
    print(f"  DARE DATASET EXPLORATION")
    print(f"  var_key    : {var_key_path.resolve()}")
    print(f"  measurands : {measurands_path.resolve()}")
    print(f"  dare_root  : "
          f"{dare_root.resolve() if dare_root else '(not provided -- skipping live scan)'}")
    print(f"  output_dir : {output_dir.resolve()}")
    print(f"{'=' * 80}")

    # Load reference files
    print("\nLoading reference files...")
    var_rows = load_variable_key(var_key_path)
    print(f"  variable_name_key.csv  : {len(var_rows)} features")
    meas = load_measurands(measurands_path)
    print(f"  Measurands XLSX        : {len(meas['utg'])} UTG rows, "
          f"{len(meas['net'])} network rows")

    # Optional live scan
    dare_stats = {}
    if dare_root:
        print(f"\nScanning {dare_root} ...")
        print("  (Reads folder names only -- no XLSX files opened)")
        dare_stats = scan_dare_root(dare_root)
        total = sum(c["on"] + c["off"] for c in dare_stats.values())
        print(f"  Found {len(dare_stats)} sessions, {total} run folders")

    # Run all 8 reports
    report_01(var_rows, meas, dare_stats, output_dir)
    report_02(meas, output_dir)
    report_03(var_rows, meas, output_dir)
    report_04(var_rows, meas, output_dir)
    report_05(var_rows, meas, output_dir)
    report_06(output_dir)
    report_07(var_rows, output_dir)
    report_08(output_dir)

    # Summary
    section("EXPLORATION COMPLETE")
    reports = sorted(output_dir.iterdir())
    print(f"\n  {len(reports)} files written to {output_dir.resolve()}\n")
    for rp in reports:
        print(f"    {rp.name:<45}  {rp.stat().st_size // 1024:>4} KB")

    print("""
  Recommended reading order:
    01  Understand scale and experiment design
    04  Learn the LTE layer structure before touching feature names
    06  Memorise the 4 benchmark values -- your pipeline correctness test
    07  Study your 5 domain KPIs -- the unique telecom-expert contribution
    08  Review the 8 DQ rules before writing Great Expectations code
    03  Open 03_feature_map.csv in PyCharm or Excel for x1-x424 lookup
""")


if __name__ == "__main__":
    main()
